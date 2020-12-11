from openpyxl import load_workbook
from .models import Localidades,Empresas,Equipo
import sys
import json

def handle_uploaded_file(f):
    wb = load_workbook(filename = f)    
    print(wb.sheetnames)
    ws = wb['Hoja1']
    for row in ws.iter_rows(min_row=2, max_col=28):
        
        q = Localidades(nombre=row[13].value)
        try:            
            q.save()
        except:
            print("localidad creada 1")
        
        q = Localidades(nombre=row[23].value)
        try:            
            q.save()
        except:
            print("localidad creada 2")        
        
        q = Empresas(nombre=row[10].value)
        try:            
            q.save()
        except:
            print("Empresa creada")        

        try:         
            src_localidad = Localidades.objects.filter(nombre=row[13].value)[0]
            src_localidad_contable = Localidades.objects.filter(nombre=row[23].value)[0]
            src_empresa = Empresas.objects.filter(nombre=row[10].value)[0]
            tl = json.dumps({'telefono 1': row[20].value, 'telefono 2': row[11].value, 'telefono 3': row[25].value,'telefono 4': row[26].value})
            q = Equipo(codm_barras = row[0].value,codm = row[1].value,id_c = int(row[2].value),nro_serie_barras = row[3].value,nro_serie = row[4].value,nro_tarea_barras = row[5].value,nro_tarea = int(row[6].value),nro_cuenta = int(row[7].value),nya_cliente = row[8].value or "" ,id_actividad = row[9].value or "",empresa = src_empresa,telefonos = tl,domicilio = row[12].value or "" ,localidad = src_localidad,codigo_postal = row[14].value or "" ,item_tarea = row[15].value or "" ,fecha_envio = row[16].value,fecha_creacion = row[17].value or row[16].value,zona = row[18].value or "" ,proveedor = row[19].value or "" ,nombre_cartera = row[21].value or "" ,notas = row[22].value or "" ,localidad_contable = src_localidad_contable,correo =row[24].value or "" ,email = row[27].value or "" )
            q.save()
        except:
                print("Some fk error")
    
        print("----------------------------------------------------------------")
